import requests
import openpyxl
import xlrd
from data_base.db_model import ThreatBase, DeviceTypeBase, DeviceTypeThreatBase
from data_base.db_controller import get_all_device_type, create_some


def download_ubi_file():
    url = 'https://bdu.fstec.ru/files/documents/thrlist.xlsx'
    response = requests.get(url, verify=False)
    file_Path = '../files/UBI.xlsx'

    if response.status_code == 200:
        with open(file_Path, 'wb') as file:
            file.write(response.content)
        print('File downloaded successfully')
    else:
        print('Failed to download file')


def parse_device_types(engine):
    file_Path = '../files/UBI.xlsx'
    book = openpyxl.load_workbook(file_Path)
    worksheet = book.active
    max_row = worksheet.max_row
    workbook = xlrd.open_workbook(file_Path)
    worksheet = workbook.sheet_by_index(0)
    dev_types = []
    exception = ['информационная система, иммигрированная в облако', 'облачная система',
                 'виртуальные устройства хранения, обработки и передачи данных',
                 'системное программное обеспечение, использующее реестр', 'реестр',
                 'облачная инфраструктура, созданная с использованием технологий виртуализации',
                 'технические средства воздушного кондиционирования, включая трубопроводные системы для циркуляции охлаждённого воздуха в цод, программируемые логические контроллеры, распределённые системы контроля, управленческие системы и другие программные средства контроля',
                 'система управления доступом, встроенная в операционную систему компьютера (программное обеспечение)',
                 'мобильное устройство и запущенные на  нем приложения (программное обеспечение, аппаратное устройство)',
                 'мобильные устройства (аппаратное устройство, программное обеспечение)',
                 'программное обеспечение (программы), использующее машинное обучение', 'модели машинного обучения',
                 'обучающие данные машинного обучения',
                 'программное обеспечение (программы), реализующие технологии искусственного интеллекта',
                 'информация, хранящаяся на компьютере во временных файлах (программное обеспечение)']
    for i in range(2, max_row):
        value = str(worksheet.cell_value(i, 4)).lower()
        value = value.strip('\n').strip(' ')
        if value not in exception:
            value = value.split(',')
            for j in value:
                j = j.strip(' ')
                dev_types.append(j.lower())
    for i in exception:
        dev_types.append(i)

    dev_types = set(dev_types)
    dev_types = list(dev_types)
    dev_types.sort()
    for device in dev_types:
        if device != '':
            device = device[0].upper() + device[1:]
            create_some(DeviceTypeBase(name=device), engine=engine)


def parse_ubi_threats(engine):
    file_Path = '../files/UBI.xlsx'
    book = openpyxl.load_workbook(file_Path)
    worksheet = book.active
    max_row = worksheet.max_row
    workbook = xlrd.open_workbook(file_Path)
    worksheet = workbook.sheet_by_index(0)
    threats = []
    intruder_list = {'Внешний нарушитель с низким потенциалом': (1, 1),
                     'Внешний нарушитель со средним потенциалом': (1, 2),
                     'Внешний нарушитель с высоким потенциалом': (1, 3),
                     'Внутренний нарушитель с низким потенциалом': (2, 1),
                     'Внутренний нарушитель со средним потенциалом': (2, 2),
                     'Внутренний нарушитель с высоким потенциалом': (2, 3)}

    for i in range(2, max_row):
        intruder = worksheet.cell_value(i, 3)
        external = 0
        internal = 0
        if intruder != '' and not None:
            if ';' in intruder:
                intruder = intruder.split('; ')
                for part in intruder:
                    result = intruder_list.get(part)
                    if result[0] == 1:
                        external = result[1]
                    if result[0] == 2:
                        internal = result[1]
            else:
                result = intruder_list.get(intruder)
                if result[0] == 1:
                    external = result[1]
                if result[0] == 2:
                    internal = result[1]
        code = int(worksheet.cell_value(i, 0))
        threat = ThreatBase(id=code, name=worksheet.cell_value(i, 1),
                            description=worksheet.cell_value(i, 2), external_offender=external,
                            internal_offender=internal,
                            confidentiality=worksheet.cell_value(i, 5), integrity=worksheet.cell_value(i, 6),
                            availability=worksheet.cell_value(i, 7))
        threats.append(threat)
    for i in threats:
        create_some(i, engine)


def connect_treat_device_type(engine):
    device_types = get_all_device_type(engine)
    file_Path = '../files/UBI.xlsx'
    book = openpyxl.load_workbook(file_Path)
    worksheet = book.active
    max_row = worksheet.max_row
    workbook = xlrd.open_workbook(file_Path)
    worksheet = workbook.sheet_by_index(0)
    device_type_threat = []
    for i in range(2, max_row):
        for j in device_types:
            if j.name.lower() in worksheet.cell_value(i, 4).lower():
                device_type_threat.append(
                    DeviceTypeThreatBase(threat_id=int(worksheet.cell_value(i, 0)), device_type_id=j.id))
    for i in device_type_threat:
        create_some(i, engine)
