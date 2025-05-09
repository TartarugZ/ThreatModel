import requests
import openpyxl
import xlrd
from data_base.db_model import VulnerabilityBase, VulnerabilityTypeBase, VulnerabilitySoftwareBase, \
    VulnerabilityTypeVulnerabilityBase, VulnerabilitySoftwareTypeBase
from data_base.db_controller import get_all_vulnerabilities_type, get_all_software_type, get_vulnerability_by_code, create_some


def parse_software_types_and_vul_types(engine):
    file_Path = '../files/Vul.xlsx'
    book = openpyxl.load_workbook(file_Path)
    worksheet = book.active
    max_row = worksheet.max_row
    workbook = xlrd.open_workbook(file_Path)
    worksheet = workbook.sheet_by_index(0)
    software_types = ''
    vulnerability_types = ''
    for i in range(3, max_row):
        soft = str(worksheet.cell_value(i, 4))
        while '(' in soft and ')' in soft:
            n = soft.index('(')
            m = soft.index(')')
            if n > m:
                break
            if m == len(soft) - 1:
                soft = soft[:n]
            else:
                soft = soft[:n] + soft[m + 1:]
        if ', ' in soft:
            soft.replace(', ', ',')
        software_types += f'{str(soft)},'
        vulnerability_types += f'{str(worksheet.cell_value(i, 8))},'
    software_types = software_types.split(',')
    vulnerability_types = vulnerability_types.split(',')
    software_types = set(software_types)
    software_types = list(software_types)
    temp = 0
    for i in range(len(software_types)):
        if software_types[i] == '':
            temp += 1
        software_types[i] = software_types[i].strip(' ')
    for i in range(temp):
        software_types.remove('')
    software_types = set(software_types)
    software_types = list(software_types)
    software_types.sort()
    vulnerability_types = set(vulnerability_types)
    vulnerability_types = list(vulnerability_types)
    temp = 0
    for i in vulnerability_types:
        if i == '':
            temp += 1
    for i in range(temp):
        vulnerability_types.remove('')
    vulnerability_types.sort()
    for software in software_types:
        create_some(VulnerabilitySoftwareBase(name=software), engine=engine)
    for vul_type in vulnerability_types:
        create_some(VulnerabilityTypeBase(name=vul_type), engine=engine)


def download_vul_file():
    url = 'https://bdu.fstec.ru/files/documents/vullist.xlsx'
    response = requests.get(url, verify=False)
    file_Path = '../files/Vul.xlsx'
    if response.status_code == 200:
        with open(file_Path, 'wb') as file:
            file.write(response.content)
        print('File downloaded successfully')
    else:
        print('Failed to download file')


def parse_vulnerabilities(engine):
    file_Path = '../files/Vul.xlsx'
    book = openpyxl.load_workbook(file_Path)
    worksheet = book.active
    max_row = worksheet.max_row
    workbook = xlrd.open_workbook(file_Path)
    worksheet = workbook.sheet_by_index(0)
    vulnerabilities = []
    for i in range(3, max_row):
        description = str(worksheet.cell_value(i, 2)) + '\n' + str(worksheet.cell_value(i, 3)) + '\n' + str(
            worksheet.cell_value(i, 5)) + '\n' + str(worksheet.cell_value(i, 6)) + '\n' + str(
            worksheet.cell_value(i, 7)) + '\n' + str(
            worksheet.cell_value(i, 12)) + '\n' + str(worksheet.cell_value(i, 13)) + '\n' + str(
            worksheet.cell_value(i, 14)) + '\n' + str(
            worksheet.cell_value(i, 16)) + '\n' + str(worksheet.cell_value(i, 18)) + '\n' + str(
            worksheet.cell_value(i, 19)) + '\n' + 'Описание угрозы: ' + str(
            worksheet.cell_value(i, 22)) + '\n' + 'Описание ошибки: ' + str(worksheet.cell_value(i, 23)) + '\n' + str(
            worksheet.cell_value(i, 24))
        vulnerabilities.append(VulnerabilityBase(code=worksheet.cell_value(i, 0), name=worksheet.cell_value(i, 1),
                                                 description=description))
    for vul in vulnerabilities:
        create_some(vul, engine)


def connect_vul_soft_vul_type(engine):
    file_Path = '../files/Vul.xlsx'
    software_type = get_all_software_type(engine)
    vul_type = get_all_vulnerabilities_type(engine)
    book = openpyxl.load_workbook(file_Path)
    worksheet = book.active
    max_row = worksheet.max_row
    workbook = xlrd.open_workbook(file_Path)
    worksheet = workbook.sheet_by_index(0)
    vul_vul_type = []
    vul_soft_type = []
    for i in range(3, max_row):
        vulnerability_id = get_vulnerability_by_code(engine, worksheet.cell_value(i, 0)).id
        for j in vul_type:
            if j.name.lower() in worksheet.cell_value(i, 8).lower():
                vul_vul_type.append(
                    VulnerabilityTypeVulnerabilityBase(vulnerability_id=vulnerability_id, vulnerability_type_id=j.id))
        for j in software_type:
            if j.name.lower() in worksheet.cell_value(i, 4).lower():
                vul_soft_type.append(
                    VulnerabilitySoftwareTypeBase(
                        vulnerability_id=vulnerability_id, software_type_id=j.id))
    for i in vul_vul_type:
        create_some(i, engine)
    for i in vul_soft_type:
        create_some(i, engine)
